import requests
import tabula
import openpyxl
import bs4

# downloadUrl = "http://treasury.ge/common/get_doc.aspx?doc_id=20035"

# OVerwrites coma seperated numbers with real numbers and copies values to main file
def changeExcel(excelName, mainFile):
    wb = openpyxl.load_workbook(excelName)
    ws = wb.active
    wbMain = openpyxl.load_workbook(mainFile)
    wsMain = wbMain.active
    x = 2
    while (wsMain.cell(2, x).value != None):
        print("in while loop" + str(x))
        x += 1
    for i in range(2,17):
        ws['C' + str(i)].value = str(ws['C' + str(i)].value).replace(",","")
        ws['D' + str(i)].value = str(ws['D' + str(i)].value).replace(",","")
        wsMain.cell(i, x).value = ws['D' + str(i)].value
    wb.save("Treasury.xlsx")    
    wb.close()
    wbMain.save(mainFile)
    wbMain.close()

# Converts pdf file to excel
def convertToExcel(filename):
    table = tabula.read_pdf(filename, pages = 1, lattice = True)[0]
    table.columns = table.columns.str.replace('\r', ' ')
    clearTable = table.dropna()
    excelName = 'Treasury.xlsx'
    clearTable.to_excel(excelName)


# Downloads pdf file from treasury.ge/5302
def downloadFile():
    page = requests.get('http://treasury.ge/5302')
    xml = bs4.BeautifulSoup(page.content, 'html.parser')
    pdfs = []
    for doc in xml.find_all('a', class_='PDF'):
        pdfs.append(doc)
    neededDoc = pdfs[2]
    href = neededDoc.get('href')
    finalURL = 'http://treasury.ge' + href
    req = requests.get(finalURL, allow_redirects = True)
    open('D:\GitHub\Treasury.pdf', 'wb').write(req.content)

def main():
    downloadFile()
    filename = "Treasury.pdf"
    convertToExcel(filename)
    changeExcel("D:\GitHub\Treasury.xlsx", "D:\GitHub\Main.xlsx")

main()