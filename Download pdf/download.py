
import requests
import tabula
import openpyxl
import bs4

downloadUrl = "http://treasury.ge/common/get_doc.aspx?doc_id=20035"

excelName = "Treasury.xlsx"

def changeExcel(excelName):
    wb = openpyxl.load_workbook(excelName)
    ws = wb.active
    values = []
    for i in range(2,17):
        ws['C' + str(i)].value = str(ws['C' + str(i)].value).replace(",","")
        ws['D' + str(i)].value = str(ws['D' + str(i)].value).replace(",","")
        values.append(ws['D' + str(i)].value)
    wb.save("Treasury.xlsx")
    wb.close()
    return values


def convertToExcel(filename):
    table = tabula.read_pdf(filename, pages = 1, lattice = True)[0]
    table.columns = table.columns.str.replace('\r', ' ')
    clearTable = table.dropna()
    excelName = 'Treasury.xlsx'
    clearTable.to_excel(excelName)


def copyToMainExcel(values, mainFile):
    wb = openpyxl.load_workbook(mainFile)
    ws = wb.active
    col = 2
    while (ws.cell(2, col).value != None):
        col += 1
        print("In while loop")
    for row in range (2,17):
        ws.cell(row, col).value = values[row - 2]
    wb.save(mainFile)
    wb.close()


def downloadFile(url, filename=''):
    try:
        if filename:
            pass            
        else:
            filename = req.url[downloadUrl.rfind('/')+1:]
        with requests.get(url) as req:
            with open(filename, 'wb') as f:
                for chunk in req.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            return filename
    except Exception as e:
        print(e)
        return None


def downloadFile2():
    page = requests.get('http://treasury.ge/5302')
    xml = bs4.BeautifulSoup(page.content, 'html.parser')
    pdfs = []
    for doc in xml.find_all('a', class_='PDF'):
        pdfs.append(doc)
    neededDoc = pdfs[2]
    href = neededDoc.get('href')
    finalURL = 'http://treasury.ge' + href
    req = requests.get(finalURL, allow_redirects = True)
    open('D:\GitHub\Treasury.pdf', 'wb').write(req.content)



def main():
    filename = "Treasury.pdf"

    downloadFile(downloadUrl, filename)
    #downloadFile(downloadUrl, filename)
    convertToExcel(filename)

    values = changeExcel("D:\GitHub\Treasury.xlsx")
        

    copyToMainExcel(values,"D:\GitHub\Main.xlsx")

main()